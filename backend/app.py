from flask import Flask, request, jsonify, send_file, send_from_directory, session, redirect, url_for
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import os
import stripe
stripe.api_key = os.environ.get("STRIPE_SECRET_KEY", "")
STRIPE_PRICE_ID = os.environ.get("STRIPE_PRICE_ID", "")
STRIPE_PK = os.environ.get("STRIPE_PK", "")
from extractor import extraer_datos_factura, guardar_en_excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import db, Usuario, InvitadoIP

app = Flask(__name__, static_folder=".")
app.secret_key = "factura-ai-clave-secreta-2024"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///usuarios.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

CORS(app)
db.init_app(app)

UPLOAD_FOLDER = "uploads"
EXCEL_PATH = "resultados.xlsx"
LIMITE_GRATIS_INVITADO = 10
LIMITE_GRATIS_REGISTRADO = 30

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Crea las tablas al arrancar
with app.app_context():
    db.create_all()

@app.route("/landing")
def landing():
    return send_from_directory(".", "landing.html")

# ── PÁGINA PRINCIPAL ──────────────────────────────────
@app.route("/")
def inicio():
    if "usuario_id" not in session:
        return send_from_directory(".", "landing.html")
    return send_from_directory(".", "index.html")

@app.route("/app")
def app_principal():
    return send_from_directory(".", "index.html")


# ── REGISTRO ──────────────────────────────────────────
@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "GET":
        return send_from_directory(".", "registro.html")
    datos = request.get_json()
    email = datos.get("email", "").strip().lower()
    contraseña = datos.get("contraseña", "")
    if not email or not contraseña:
        return jsonify({"error": "Email y contraseña son obligatorios"}), 400
    if Usuario.query.filter_by(email=email).first():
        return jsonify({"error": "Este email ya está registrado"}), 400
    facturas_invitado = session.get("facturas_invitado", 0)
    nuevo = Usuario(email=email, contraseña=generate_password_hash(contraseña), facturas_usadas=facturas_invitado)
    db.session.add(nuevo)
    db.session.commit()
    session["usuario_id"] = nuevo.id
    session["usuario_email"] = nuevo.email
    return jsonify({"ok": True})


# ── LOGIN ─────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return send_from_directory(".", "login.html")
    datos = request.get_json()
    email = datos.get("email", "").strip().lower()
    contraseña = datos.get("contraseña", "")
    usuario = Usuario.query.filter_by(email=email).first()
    if not usuario or not check_password_hash(usuario.contraseña, contraseña):
        return jsonify({"error": "Email o contraseña incorrectos"}), 401
    facturas_invitado = session.get("facturas_invitado", 0)
    if facturas_invitado > 0:
        usuario.facturas_usadas = max(usuario.facturas_usadas, facturas_invitado)
        db.session.commit()
    session["usuario_id"] = usuario.id
    session["usuario_email"] = usuario.email
    return jsonify({"ok": True})


# ── LOGOUT ────────────────────────────────────────────
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ── PROCESAR FACTURAS ─────────────────────────────────
@app.route("/procesar", methods=["POST"])
def procesar_facturas():
    if "usuario_id" not in session:
        ip = request.remote_addr
        invitado = InvitadoIP.query.filter_by(ip=ip).first()
        if not invitado:
            invitado = InvitadoIP(ip=ip, facturas_usadas=0)
            db.session.add(invitado)
            db.session.commit()
        if invitado.facturas_usadas >= LIMITE_GRATIS_INVITADO:
            return jsonify({"error": "limite_invitado"}), 403
    else:
        usuario = Usuario.query.get(session["usuario_id"])
        if not usuario.es_premium and usuario.facturas_usadas >= LIMITE_GRATIS_REGISTRADO:
            return jsonify({"error": "limite_alcanzado"}), 403

    archivos = request.files.getlist("archivos")
    if not archivos:
        return jsonify({"error": "No se han enviado archivos"}), 400

    excel_base = None
    if "excel_base" in request.files:
        archivo_excel = request.files["excel_base"]
        if archivo_excel.filename.endswith(".xlsx"):
            excel_base = os.path.join(UPLOAD_FOLDER, "base_" + archivo_excel.filename)
            archivo_excel.save(excel_base)

    extensiones_permitidas = {".pdf"}
    resultados = []

    if os.path.exists(EXCEL_PATH):
        os.remove(EXCEL_PATH)

    for archivo in archivos:
        extension = os.path.splitext(archivo.filename)[1].lower()
        if extension not in extensiones_permitidas:
            resultados.append({"archivo": archivo.filename, "error": "Formato no permitido"})
            continue
        ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
        archivo.save(ruta)
        datos = extraer_datos_factura(ruta)
        guardar_en_excel(datos, EXCEL_PATH, excel_base)

        if "usuario_id" not in session:
            invitado.facturas_usadas += 1
            db.session.commit()
        else:
            usuario.facturas_usadas += 1
            db.session.commit()

        resultados.append({"archivo": archivo.filename, "datos": datos})

    if "usuario_id" not in session:
        return jsonify({
            "resultados": resultados,
            "facturas_usadas": invitado.facturas_usadas,
            "es_premium": False,
            "es_invitado": True,
            "facturas_restantes": max(0, LIMITE_GRATIS_INVITADO - invitado.facturas_usadas)
        })
    else:
        return jsonify({
            "resultados": resultados,
            "facturas_usadas": usuario.facturas_usadas,
            "es_premium": usuario.es_premium,
            "es_invitado": False,
            "facturas_restantes": max(0, LIMITE_GRATIS_REGISTRADO - usuario.facturas_usadas) if not usuario.es_premium else "ilimitadas"
        })


# ── DESCARGAR EXCEL ───────────────────────────────────
@app.route("/descargar", methods=["GET"])
def descargar_excel():
    if "usuario_id" not in session:
        return jsonify({"error": "Debes iniciar sesión"}), 401
    if os.path.exists(EXCEL_PATH):
        return send_file(EXCEL_PATH, as_attachment=True)
    return jsonify({"error": "No hay Excel generado todavía"}), 404


# ── ESTADO DEL USUARIO ────────────────────────────────
@app.route("/estado")
def estado_usuario():
    if "usuario_id" not in session:
        ip = request.remote_addr
        invitado = InvitadoIP.query.filter_by(ip=ip).first()
        facturas_usadas = invitado.facturas_usadas if invitado else 0
        return jsonify({
            "autenticado": False,
            "es_invitado": True,
            "facturas_usadas": facturas_usadas,
            "facturas_restantes": max(0, LIMITE_GRATIS_INVITADO - facturas_usadas)
        })
    usuario = Usuario.query.get(session["usuario_id"])
    return jsonify({
        "autenticado": True,
        "email": usuario.email,
        "facturas_usadas": usuario.facturas_usadas,
        "es_premium": usuario.es_premium,
        "es_invitado": False,
        "facturas_restantes": max(0, LIMITE_GRATIS_REGISTRADO - usuario.facturas_usadas) if not usuario.es_premium else "ilimitadas"
    })


# ── CREAR SESIÓN DE PAGO ──────────────────────────────
@app.route("/crear-pago", methods=["POST"])
def crear_pago():
    if "usuario_id" not in session:
        return jsonify({"error": "Debes iniciar sesión"}), 401

    if not stripe.api_key:
        return jsonify({"error": "Stripe no configurado"}), 500

    usuario = Usuario.query.get(session["usuario_id"])

    # URL dinámica: funciona en local Y en producción automáticamente
    base_url = request.host_url.rstrip("/")

    checkout = stripe.checkout.Session.create(
        payment_method_types=["card"],
        mode="subscription",
        line_items=[{"price": STRIPE_PRICE_ID, "quantity": 1}],
        customer_email=usuario.email,
        success_url=f"{base_url}/pago-exitoso?session_id={{CHECKOUT_SESSION_ID}}",
        cancel_url=f"{base_url}/",
    )
    return jsonify({"url": checkout.url})


# ── PAGO EXITOSO ──────────────────────────────────────
@app.route("/pago-exitoso")
def pago_exitoso():
    if "usuario_id" not in session:
        return redirect("/login")
    usuario = Usuario.query.get(session["usuario_id"])
    usuario.es_premium = True
    db.session.commit()
    return redirect("/")


# ── CLAVE PÚBLICA STRIPE ──────────────────────────────
@app.route("/stripe-pk")
def stripe_pk():
    return jsonify({"pk": STRIPE_PK})


# ── CUENTA ────────────────────────────────────────────
@app.route("/cuenta")
def cuenta():
    if "usuario_id" not in session:
        return redirect("/login")
    return send_from_directory(".", "cuenta.html")

@app.route("/cuenta-datos")
def cuenta_datos():
    if "usuario_id" not in session:
        return jsonify({"error": "No autenticado"}), 401
    usuario = Usuario.query.get(session["usuario_id"])
    return jsonify({
        "email": usuario.email,
        "es_premium": usuario.es_premium,
        "facturas_usadas": usuario.facturas_usadas
    })

@app.route("/portal-cliente")
def portal_cliente():
    if "usuario_id" not in session:
        return redirect("/login")
    usuario = Usuario.query.get(session["usuario_id"])
    try:
        clientes = stripe.Customer.list(email=usuario.email, limit=1)
        if not clientes.data:
            return redirect("/cuenta?error=no_cliente")
        customer_id = clientes.data[0].id
        base_url = request.headers.get("Origin") or request.host_url.rstrip("/")
        if "onrender.com" in base_url and base_url.startswith("http://"):
            base_url = base_url.replace("http://", "https://")
        portal = stripe.billing_portal.Session.create(
            customer=customer_id,
            return_url=f"{base_url}/cuenta"
        )
        return redirect(portal.url)
    except Exception as e:
        return redirect("/cuenta?error=portal")

@app.route("/leer-excel", methods=["POST"])
def leer_excel():
    if "usuario_id" not in session:
        return jsonify({"error": "Debes iniciar sesión"}), 401

    if "excel" not in request.files:
        if os.path.exists(EXCEL_PATH):
            libro = openpyxl.load_workbook(EXCEL_PATH)
        else:
            return jsonify({"filas": [], "cabeceras": []})
    else:
        archivo = request.files["excel"]
        ruta = os.path.join(UPLOAD_FOLDER, "editor_" + archivo.filename)
        archivo.save(ruta)
        libro = openpyxl.load_workbook(ruta)

    hoja = libro.active
    filas = []
    for row in hoja.iter_rows(values_only=True):
        filas.append(list(row))

    cabeceras = filas[0] if filas else []
    datos = filas[1:] if len(filas) > 1 else []

    return jsonify({"cabeceras": cabeceras, "filas": datos})


@app.route("/guardar-excel", methods=["POST"])
def guardar_excel():
    if "usuario_id" not in session:
        return jsonify({"error": "Debes iniciar sesión"}), 401

    datos = request.get_json()
    cabeceras = datos.get("cabeceras", [])
    filas = datos.get("filas", [])

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Facturas"

    verde_oscuro = PatternFill("solid", fgColor="1A3C34")
    verde_claro  = PatternFill("solid", fgColor="D6EAE4")
    blanco       = PatternFill("solid", fgColor="FFFFFF")
    fuente_cab   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal = Font(name="Calibri", size=10)
    borde_lado = Side(style="thin", color="B0C4BE")
    borde = Border(left=borde_lado, right=borde_lado, top=borde_lado, bottom=borde_lado)

    for col, titulo in enumerate(cabeceras, start=1):
        celda = hoja.cell(row=1, column=col, value=titulo)
        celda.font = fuente_cab
        celda.fill = verde_oscuro
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
    hoja.row_dimensions[1].height = 22

    for i, fila in enumerate(filas, start=2):
        relleno = verde_claro if i % 2 == 0 else blanco
        for col, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=i, column=col, value=valor)
            celda.font = fuente_normal
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde
        hoja.row_dimensions[i].height = 18

    for col in range(1, len(cabeceras) + 1):
        hoja.column_dimensions[get_column_letter(col)].width = 22

    ruta_nueva = os.path.join(UPLOAD_FOLDER, "editado.xlsx")
    libro.save(ruta_nueva)
    return send_file(ruta_nueva, as_attachment=True, download_name="facturas_editadas.xlsx")

@app.route("/test-ocr", methods=["POST"])
def test_ocr():
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"error": "No hay archivo"})
    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)
    from extractor import extraer_texto
    texto = extraer_texto(ruta)
    return jsonify({"texto": texto})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)