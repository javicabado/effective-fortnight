import pdfplumber
import re
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def extraer_texto(ruta_archivo):
    extension = os.path.splitext(ruta_archivo)[1].lower()
    if extension == ".pdf":
        texto = ""
        with pdfplumber.open(ruta_archivo) as pdf:
            for pagina in pdf.pages:
                texto += pagina.extract_text() or ""
        return texto
    return ""


def extraer_con_ia(texto):
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return None
    try:
        from groq import Groq
        cliente = Groq(api_key=api_key)
        prompt = f"""Eres un experto en contabilidad española. Analiza este texto de una factura y extrae exactamente estos 7 campos.

REGLAS CRÍTICAS:
- "Razón Social" = el EMISOR (quien COBRA, el proveedor), NO el cliente
- "CIF" = el NIF/CIF del EMISOR, no del cliente
- "Base Imponible", "IVA", "Total" = solo el número con 2 decimales, sin € ni texto. Ejemplo: 1234.56
- "Fecha" = formato DD/MM/YYYY si es posible
- Si un campo no existe en el texto escribe exactamente: No encontrado
- Responde ÚNICAMENTE con el JSON, sin texto antes ni después, sin bloques de código

JSON requerido:
{{"Razón Social": "...", "CIF": "...", "Número Factura": "...", "Fecha": "...", "Base Imponible": "...", "IVA": "...", "Total": "..."}}

TEXTO DE LA FACTURA:
{texto[:4000]}"""

        respuesta = cliente.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )
        texto_respuesta = respuesta.choices[0].message.content.strip()
        texto_respuesta = re.sub(r'^```(?:json)?\s*', '', texto_respuesta)
        texto_respuesta = re.sub(r'\s*```$', '', texto_respuesta)
        match = re.search(r'\{.*\}', texto_respuesta, re.DOTALL)
        if match:
            texto_respuesta = match.group(0)
        datos = json.loads(texto_respuesta)
        campos = ["Razón Social", "CIF", "Número Factura", "Fecha", "Base Imponible", "IVA", "Total"]
        for campo in campos:
            if campo not in datos:
                datos[campo] = "No encontrado"
        for campo in ["Base Imponible", "IVA", "Total"]:
            valor = str(datos.get(campo, "No encontrado")).strip()
            if valor and valor != "No encontrado":
                valor = valor.replace(",", ".").replace("€", "").replace("EUR", "").strip()
                if re.match(r'^\d+(\.\d{1,2})?$', valor):
                    datos[campo] = valor + " EUR"
                else:
                    datos[campo] = valor
        return datos
    except Exception:
        return None


def extraer_con_regex(texto_crudo):
    datos = {k: "No encontrado" for k in
             ["Razón Social", "CIF", "Número Factura", "Fecha", "Base Imponible", "IVA", "Total"]}

    lineas = [l.strip() for l in texto_crudo.strip().split("\n") if l.strip()]
    texto = " ".join(lineas)
    sep_cliente = r'(?:\n[Ff]actura[:\s]+[A-Z0-9][A-Z0-9\-\/]{2,25}\n|FACTURAR?\s+A\b|FACTURADO\s+A\b)'
    partes = re.split(sep_cliente, texto_crudo, maxsplit=1, flags=re.IGNORECASE)
    texto_emisor = partes[0] if len(partes) > 1 else texto_crudo
    lineas_emisor = [l.strip() for l in texto_emisor.split("\n") if l.strip()]

    razon_encontrada = False
    SUFIJO = r'(?:S\.?[ \t]?L\.?[ \t]?U?\.?|S\.?[ \t]?A\.?[ \t]?U?\.?|S\.?[ \t]?C\.?|S\.?[ \t]?L\.?[ \t]?P\.|,[ \t]?S\.?[ \t]?[LA]\.?)'

    m = re.search(r'DATOS\s+DEL\s+EMISOR[:\s\n]*([^\n]+)', texto_crudo, re.IGNORECASE)
    if m:
        datos["Razón Social"] = m.group(1).strip()[:60]
        razon_encontrada = True

    if not razon_encontrada:
        patron_2 = r'([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñÁÉÍÓÚÑ \t&\.,]{2,40}' + SUFIJO + r')[^\n]{0,40}(?:CIF|NIF)'
        for tb in [texto_emisor, texto_crudo]:
            m = re.search(patron_2, tb, re.IGNORECASE)
            if m:
                break
        if m:
            valor = m.group(1).strip().rstrip(".,; \t")
            if len(valor) > 3:
                datos["Razón Social"] = valor[:60]
                razon_encontrada = True

    if not razon_encontrada:
        m = re.search(r'([A-ZÁÉÍÓÚÑ][A-Za-záéíóúñÁÉÍÓÚÑ\s&\.,]{2,50}' + SUFIJO + r')', texto_emisor, re.IGNORECASE)
        if m:
            valor = m.group(1).strip()
            m2 = re.search(SUFIJO, valor, re.IGNORECASE)
            if m2: valor = valor[:m2.end()].strip()
            if len(valor) > 3:
                datos["Razón Social"] = valor[:60]
                razon_encontrada = True

    if not razon_encontrada:
        ruido = [r'^\d+$', r'^factura$', r'^(?:CIF|NIF|DNI)', r'^AÑO', r'^MES\b', r'^DÍA', r'^DOC', r'^ref\.']
        for linea in lineas_emisor[:10]:
            if not any(re.search(p, linea, re.IGNORECASE) for p in ruido) and len(linea) > 4 and re.search(r'[A-Za-záéíóúñ]{3}', linea):
                datos["Razón Social"] = linea[:60]
                break

    patrones_cif = [
        r'(?:CIF|NIF|N\.I\.F|C\.I\.F|DNI)[:\s\.\-]*([A-Za-z]-?\d{7}-?[A-Za-z0-9])',
        r'(?:CIF|NIF|N\.I\.F|C\.I\.F|DNI)[:\s\.\-]*([0-9]{8}[A-Za-z])',
        r'\b([A-HJ-NP-SUVW]-?\d{7}-?[0-9A-J])\b',
        r'\b([0-9]{8}[A-Za-z])\b',
        r'\b([XYZ]-?\d{7}[A-Za-z])\b',
    ]
    for patron in patrones_cif[:2]:
        m = re.search(patron, texto_emisor, re.IGNORECASE)
        if m: datos["CIF"] = m.group(1).upper(); break
    if datos["CIF"] == "No encontrado":
        for patron in patrones_cif:
            m = re.search(patron, texto_crudo[-600:], re.IGNORECASE)
            if m: datos["CIF"] = m.group(1).upper(); break
    if datos["CIF"] == "No encontrado":
        for patron in patrones_cif:
            m = re.search(patron, texto, re.IGNORECASE)
            if m: datos["CIF"] = m.group(1).upper(); break

    for patron in [
        r'N[º°][:\s]+([A-Z0-9][A-Z0-9\-\/]{3,25})\b',
        r'CONTINUACI[OÓ]N\s*FACTURA\s+([A-Z0-9][A-Z0-9\-\/]{3,25})',
        r'(?:factura\s*n[uú]m(?:ero)?\.?|n[uú]m(?:ero)?\s*(?:de\s*)?factura)[:\s\.\-#]*([A-Z0-9][A-Z0-9\-\/]{2,20})',
        r'[Ff]actura\s+([A-Z][0-9]{4}[A-Z0-9\-\/]{1,15})\b',
        r'\b([A-Z]{1,4}[\/\-]\d{4}[\/\-][A-Z]{2,4}[\/\-]\d{4})\b',
        r'\b([A-Z]{1,4}[\/\-]\d{4}[\/\-]\d{2}[\/\-]\d{2,6})\b',
        r'\b(?:FAC|FRA|INV|F)[\/\-](\d{2,6}(?:[\/\-]\d{2,4})?)\b',
        r'[Ff]actura[:\s]*([0-9]{4}[\/\-][0-9]{2,6})',
        r'[Ff]-(\d{2,4}-?\d{2,6})',
        r'[Nn][uú]m(?:ero)?\.?\s*[:\s]*(\d{3,8})',
    ]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m: datos["Número Factura"] = m.group(1).strip(); break

    for patron in [
        r'(?:fecha\s*(?:de\s*)?(?:factura|emisi[oó]n)?)[:\s]*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})',
        r'(?:fecha)[:\s]*(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
        r'\b(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4})\b',
        r'\b(\d{1,2}\s+de\s+(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+\d{4})\b',
    ]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m: datos["Fecha"] = m.group(1).strip(); break

    NUM = r'([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{1,2})?)\s*(?:€|EUR)?'
    for patron in [r'[Bb]ase\s*imponible\s*total[:\s]*'+NUM, r'[Bb]ase\s*[Ii]mponible[:\s]*'+NUM, r'[Ii]mporte\s+base[:\s]*'+NUM, r'[Ss]ubtotal\s*(?:sin\s*impuestos|sin\s*IVA)?[:\s]*'+NUM, r'[Nn]eto[:\s]*'+NUM]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m: datos["Base Imponible"] = limpiar_numero(m.group(1)) + " EUR"; break

    for patron in [r'IVA\s*\(\s*(?:21|10|4)\s*%[^)]*\)[:\s]*'+NUM, r'IVA\s*(?:21|10|4)\s*%[:\s]*'+NUM, r'[Cc]uota\s*IVA[:\s]*'+NUM, r'\bIVA[:\s]+'+NUM]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m: datos["IVA"] = limpiar_numero(m.group(1)) + " EUR"; break

    for patron in [r'\bTOTAL\s*A\s*PAGAR\s*(?:\(EUR\))?[:\s]*'+NUM, r'\bTOTAL\s*FACTURA[:\s]*'+NUM, r'\bTOTAL\s*EUR[:\s]*'+NUM, r'\bTOTAL[^a-zA-Z0-9]{0,5}'+NUM, r'[Ii]mporte\s*(?:total|a\s*pagar)[^0-9]{0,20}'+NUM]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m: datos["Total"] = limpiar_numero(m.group(1)) + " EUR"; break

    if datos["IVA"] == "No encontrado" and datos["Base Imponible"] != "No encontrado" and datos["Total"] != "No encontrado":
        try:
            base = float(datos["Base Imponible"].replace(" EUR","").replace(",","."))
            total = float(datos["Total"].replace(" EUR","").replace(",","."))
            iva = round(total - base, 2)
            if 0 < iva < total: datos["IVA"] = str(iva) + " EUR"
        except Exception:
            pass

    return datos


def extraer_datos_factura(ruta_archivo):
    texto_crudo = extraer_texto(ruta_archivo)
    if not texto_crudo:
        return {k: "No encontrado" for k in ["Razón Social","CIF","Número Factura","Fecha","Base Imponible","IVA","Total"]}
    resultado_ia = extraer_con_ia(texto_crudo)
    if resultado_ia:
        return resultado_ia
    return extraer_con_regex(texto_crudo)


def limpiar_numero(texto):
    texto = texto.strip()
    if re.match(r'^\d{1,3}(\.\d{3})+(,\d{1,2})?$', texto):
        texto = texto.replace(".", "").replace(",", ".")
    elif re.match(r'^\d+(,\d{1,2})$', texto):
        texto = texto.replace(",", ".")
    elif re.match(r'^\d{1,3}(\.\d{3})+$', texto):
        texto = texto.replace(".", "")
    return texto


def guardar_en_excel(datos, ruta_excel="resultados.xlsx", excel_base=None):
    verde_oscuro    = PatternFill("solid", fgColor="1A3C34")
    verde_claro     = PatternFill("solid", fgColor="D6EAE4")
    blanco          = PatternFill("solid", fgColor="FFFFFF")
    fuente_cabecera = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal   = Font(name="Calibri", size=10)
    borde_lado      = Side(style="thin", color="B0C4BE")
    borde           = Border(left=borde_lado, right=borde_lado, top=borde_lado, bottom=borde_lado)
    cabeceras       = ["Razón Social", "CIF", "Número Factura", "Fecha", "Base Imponible", "IVA", "Total"]

    if excel_base and os.path.exists(excel_base):
        libro = openpyxl.load_workbook(excel_base); hoja = libro.active; es_nuevo = False
    elif os.path.exists(ruta_excel):
        libro = openpyxl.load_workbook(ruta_excel); hoja = libro.active; es_nuevo = False
    else:
        libro = openpyxl.Workbook(); hoja = libro.active; hoja.title = "Facturas"; es_nuevo = True

    if es_nuevo:
        for col, titulo in enumerate(cabeceras, start=1):
            celda = hoja.cell(row=1, column=col, value=titulo)
            celda.font = fuente_cabecera; celda.fill = verde_oscuro
            celda.alignment = Alignment(horizontal="center", vertical="center"); celda.border = borde
        hoja.row_dimensions[1].height = 22

    fila = hoja.max_row + 1
    valores = [datos["Razón Social"], datos["CIF"], datos["Número Factura"],
               datos["Fecha"], datos["Base Imponible"], datos["IVA"], datos["Total"]]
    relleno = verde_claro if fila % 2 == 0 else blanco

    for col, valor in enumerate(valores, start=1):
        celda = hoja.cell(row=fila, column=col, value=valor)
        celda.font = fuente_normal; celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center"); celda.border = borde

    hoja.row_dimensions[fila].height = 18
    for col in range(1, 8):
        hoja.column_dimensions[get_column_letter(col)].width = 22
    libro.save(ruta_excel)