import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from PIL import Image
import pytesseract

# Ruta donde Windows instala Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


def extraer_texto(ruta_archivo):
    extension = os.path.splitext(ruta_archivo)[1].lower()

    if extension == ".pdf":
        texto = ""
        with pdfplumber.open(ruta_archivo) as pdf:
            for pagina in pdf.pages:
                texto += pagina.extract_text() or ""
        return texto

    elif extension in [".jpg", ".jpeg", ".png", ".webp"]:
        imagen = Image.open(ruta_archivo)
        imagen = imagen.convert("L")  # solo escala de grises, sin binarizar
        config = "--psm 6 --oem 3"
        texto = pytesseract.image_to_string(imagen, lang="spa", config=config)
        return texto

    else:
        return ""


def extraer_datos_factura(ruta_archivo):
    datos = {
        "CIF": "No encontrado",
        "Base Imponible": "No encontrado",
        "IVA": "No encontrado",
        "Total": "No encontrado"
    }

    texto_completo = extraer_texto(ruta_archivo)

     # CIF: letra + 8 números, con o sin espacio tras "CIF:"
    cif = re.search(r'CIF[:\s]*([A-Z]\d{8})', texto_completo)
    if not cif:
        cif = re.search(r'[A-Z]\d{8}', texto_completo)
    if cif:
        datos["CIF"] = cif.group(1) if cif.lastindex else cif.group()

    # Base Imponible: acepta "Base Imponible", "Base imponible" o solo "Base"
    base = re.search(
        r'(?:Base Imponible|Base imponible|BASE IMPONIBLE)[:\s]+([0-9.,]+)',
        texto_completo
    )
    if not base:
        # Formato tabla: "IVA 21,0%   74,36   15,62   89,98"
        base = re.search(r'IVA\s+\d+[.,]\d+\s*%?\s+([0-9.,]+)', texto_completo)
    if base:
        datos["Base Imponible"] = base.group(1) + " EUR"

    # IVA: acepta "IVA (21%):", "IVA 21,0%", "Cuota"
    iva = re.search(r'IVA[^:\n]*:\s*([0-9.,]+)', texto_completo)
    if not iva:
        iva = re.search(r'IVA\s+\d+[.,]\d+\s*%?\s+[0-9.,]+\s+([0-9.,]+)', texto_completo)
    if iva:
        datos["IVA"] = iva.group(1) + " EUR"

    # Total: acepta "TOTAL FACTURA:", "TOTAL Euros:", "Total:", "TOTAL:"
    total = re.search(
        r'TOTAL[^:\n]*[:\s]+([0-9.,]+)',
        texto_completo, re.IGNORECASE
    )
    if total:
        datos["Total"] = total.group(1) + " EUR"

    return datos


def guardar_en_excel(datos, ruta_excel="resultados.xlsx"):
    verde_oscuro  = PatternFill("solid", fgColor="1A3C34")
    verde_claro   = PatternFill("solid", fgColor="D6EAE4")
    blanco        = PatternFill("solid", fgColor="FFFFFF")
    fuente_cabecera = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal   = Font(name="Calibri", size=10)
    borde_lado = Side(style="thin", color="B0C4BE")
    borde = Border(left=borde_lado, right=borde_lado,
                   top=borde_lado, bottom=borde_lado)

    cabeceras = ["CIF", "Base Imponible", "IVA", "Total"]

    if os.path.exists(ruta_excel):
        libro = openpyxl.load_workbook(ruta_excel)
        hoja  = libro.active
        es_nuevo = False
    else:
        libro    = openpyxl.Workbook()
        hoja     = libro.active
        hoja.title = "Facturas"
        es_nuevo = True

    if es_nuevo:
        for col, titulo in enumerate(cabeceras, start=1):
            celda = hoja.cell(row=1, column=col, value=titulo)
            celda.font      = fuente_cabecera
            celda.fill      = verde_oscuro
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border    = borde
        hoja.row_dimensions[1].height = 22

    fila = hoja.max_row + 1
    valores = [datos["CIF"], datos["Base Imponible"],
               datos["IVA"], datos["Total"]]
    relleno = verde_claro if fila % 2 == 0 else blanco

    for col, valor in enumerate(valores, start=1):
        celda = hoja.cell(row=fila, column=col, value=valor)
        celda.font      = fuente_normal
        celda.fill      = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border    = borde

    hoja.row_dimensions[fila].height = 18

    for col in range(1, 5):
        letra = get_column_letter(col)
        hoja.column_dimensions[letra].width = 22

    libro.save(ruta_excel)
    print(f"Datos guardados en: {ruta_excel}")


if __name__ == "__main__":
    ruta = input("Escribe la ruta del archivo (PDF o imagen): ")
    resultado = extraer_datos_factura(ruta)

    print("\n--- DATOS EXTRAÍDOS ---")
    for clave, valor in resultado.items():
        print(f"{clave}: {valor}")
    print("-----------------------")

    guardar_en_excel(resultado)